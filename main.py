import subprocess
import os
import glob
import re
from openpyxl import load_workbook
import timeit
import datetime

# Begin tracking the time it takes to run
start_time = timeit.default_timer()

WDB_JAVA_CONVERTER = "WorksDatabaseConverter.jar"

# Check if the script is running in a docker container
if "DOCKER_ENV" in os.environ:
    source_dir = "/app/source"
    output_dir = "/app/output"
    converter_path = os.environ["CONVERTER_PATH"]

    # Path to the LibreOffice executable
    libreoffice_path = os.environ["LIBREOFFICE_PATH"]
else:
    # Set Windows paths
    source_dir = os.path.expanduser("~/Downloads")
    output_dir = os.path.expanduser("~/Downloads")
    converter_path = os.path.join(source_dir, WDB_JAVA_CONVERTER)
    # Path to the Windows x64 LibreOffice executable
    libreoffice_path = r'C:\Program Files\LibreOffice\program\soffice.exe'

# Start at the first row
header_start_row = 1

# Track number of files converted
converted = 0

for root, dirs, files in os.walk(source_dir):
    for file in files:
        if file.lower().endswith(".wdb"):
            # construct the output file name by replacing the file extension with xlsx
            output_file = os.path.splitext(file)[0] + ".xlsx"

            # construct the input file path
            input_file = os.path.join(root, file)

            # construct the output file path
            output_path = os.path.join(output_dir, output_file)

            # Call the WorksDatabaseConverter.jar to get the header fields
            p = subprocess.Popen(['java', '-jar', converter_path, input_file], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            out, err = p.communicate()

            # Parse the header fields from the output
            headers = out.decode().strip().split(',')

            # Clean up the header garbage
            header_pattern = r'Header \(\d+\): ([A-Z# ]+)'

            clean_headers = []
            for header in headers:
                matches = re.findall(header_pattern, header)
                for match in matches:
                    clean_headers.append(match)

            # Call the LibreOffice CLI to convert the Works file to Excel
            subprocess.call([libreoffice_path, '--headless', '--convert-to', 'xlsx', '--outdir', output_dir, input_file])

            # Open the converted Excel file
            wb = load_workbook(output_path)
            ws = wb.active

            # Insert a new row at the beginning of the worksheet
            ws.insert_rows(header_start_row)

            # Insert the header fields into the first row of the worksheet
            for i, header in enumerate(clean_headers):
                cell = ws.cell(row=header_start_row, column=i+1)
                cell.value = header

            # Save the modified Excel file
            wb.save(output_path)

            # Remove the CSV file
            csv_file = os.path.join(source_dir, os.path.splitext(file)[0] + ".csv")
            os.remove(csv_file)

            print(f"Converted {input_file} to {output_file} in {output_dir}")
            converted += 1

runtime = timeit.default_timer() - start_time
print(f"Converted files: {converted} in {datetime.timedelta(seconds=runtime)}")
